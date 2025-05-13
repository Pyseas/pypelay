
from .geom import *
import datetime
from pathlib import Path
import OrcFxAPI as ofx
import math
import pandas as pd
import shutil
import numpy as np
import warnings
from dataclasses import dataclass, field
from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from importlib.resources import files

PATH = Path('.')

__all__ = ['fetch_files', 'static_summary', 'set_radius', 'select_radius',
           'stinger_setup', 'adjust_top_tension', 'Vessel', 'list_raos']


@dataclass
class Roller:
    name: str  # SR1, SR2 etc
    y: float = 0.0  # Roller height (in mm)
    r3: float = 0.0  # Angle in deg relative to stinger section bottom chord
    post_angle: float = 0.0  # Post angle in deg relative to vessel
    arc: float = 0.0  # Arc length along pipe path (in mm)
    y_offset: float = 0.0  # Offset from V to pipe CL (in mm)

@dataclass
class Vessel:
    """Vessel class.

    Args:
        name (str): Vessel name
        vessel_type (str): Orcaflex vessel type name
        draft_name (str): Orcaflex draft name

    Examples:
        vessel = Vessel('S1200', 'vt S1200', 'Draft_7.4m')
    """
    name: str = 'S1200'
    vessel_type: str = 'vt S1200'
    draft_name: str = 'Draft_7.4m'
    draft: float = 0.0

    def __post_init__(self):
        xlpath = Path(str(files('pypelay') / 'raos.xlsx'))
        df = pd.read_excel(xlpath)
        msg = 'not recognized, use pypelay.list_raos() to see available RAOs'

        df = df[df['Vessel'] == self.name]
        if df.empty:
            raise ValueError(f'Vessel name {msg}')

        df = df[df['Vessel type'] == self.vessel_type]
        if df.empty:
            raise ValueError(f'Vessel type {msg}')

        df = df[df['Draft name'] == self.draft_name]
        if df.empty:
            raise ValueError(f'Draft name {msg}')
        
        self.draft = float(df.iloc[0]['Draft_m'])

@dataclass
class StingerSetupArgs:
    inpath: Path
    outpath: Path
    config: dict
    # straight: float xxx
    # transition: float
    # ang1: float
    # ang2: float
    water_depth: float
    tip_clearance: float
    delete_dat: bool = False

@dataclass
class StingerSetupOptions:
    tip_def: str
    tip_len: float
    seglen: float
    ob_seglen: float
    tdp_seglen: float
    tdp_len: float
    tensioner_mode: str
    deadband: float

@dataclass
class StingerSetupResults:
    outpath: Path
    top_tension: float = -1.0
    uc_ob: float = -1.0
    uc_sag: float = -1.0
    tip_depth: float = 0.0
    tip_angle: float = 0.0
    draft: float = 0.0

    def __str__(self):
        outstr = f'{self.outpath.name} : '
        outstr += f'top_tension {self.top_tension:.1f}kN, '
        outstr += f'uc_sag {self.uc_sag:.3f}, '
        outstr += f'tip_depth {self.tip_depth:.1f}m'

        return outstr

@dataclass
class LineType:
    Name: str
    Mass: float
    OD: float
    ID: float
    CoatingThickness: float
    CoatingMaterialDensity: float
    LiningThickness: float
    LiningMaterialDensity: float
    CWC: str
    SNIF: float
    WallThickness: float = field(init=False)

    def __post_init__(self):
        self.WallThickness = (self.OD - self.ID) / 2

    def weights(self, contents_density: float) -> tuple[float, float]:
        if self.CWC == 'Yes':
            wt_in_air = self.Mass
            lining_id = self.ID
            if self.LiningThickness > 0.0:
                lining_id = self.ID - 2 * self.LiningThickness
            contents_wt = np.pi / 4 * lining_id**2 * contents_density
            disp = np.pi / 4 * self.OD**2 * 1.025
        else:
            pipe_wt = np.pi / 4 * (self.OD**2 - self.ID**2) * 7.85
            # Coating
            coating_wt = 0.0
            coating_od = self.OD
            if self.CoatingThickness > 0.0:
                coating_od = self.OD + 2 * self.CoatingThickness
                coating_id = self.OD
                coating_wt = (np.pi / 4 * (coating_od**2 - coating_id**2) *
                            self.CoatingMaterialDensity)
            # Lining
            lining_wt = 0.0
            lining_id = self.ID
            if self.LiningThickness > 0.0:
                lining_od = self.ID
                lining_id = self.ID - 2 * self.LiningThickness
                lining_wt = (np.pi / 4 * (lining_od**2 - lining_id**2) *
                            self.LiningMaterialDensity)

            wt_in_air = pipe_wt + coating_wt + lining_wt
            contents_wt = np.pi / 4 * lining_id**2 * contents_density
            disp = np.pi / 4 * coating_od**2 * 1.025

        wt_submerged = wt_in_air + contents_wt - disp

        return wt_in_air, wt_submerged


def list_raos(vessel_name: str | None = None):
    """List available RAOs for the specified vessel.
    If vessel name is not given then list all RAOs.

    Args:
        vessel_name (str): Vessel name
    """

    xlpath = Path(str(files('pypelay') / 'raos.xlsx'))
    df = pd.read_excel(xlpath)

    if vessel_name:
        df = df[df['Vessel'] == vessel_name]

    print(tabulate(df, headers='keys', tablefmt='psql', showindex=False))


def static_summary(outpath, datpaths: list[Path]):
    """Create spreadsheet with static results for one or multiple dat files.

    Args:
        outpath (Path): File path of new spreadsheet
        datpaths list[Path]: List of dat file paths
    """
    xlpath = Path(str(files('pypelay') / 'static_summary.xlsx'))
    wb = load_workbook(xlpath)
    ws = wb['Sheet1']
    # style_str = NamedStyle(name='style_str')
    # style_str.alignment = Alignment(vertical='center', horizontal='center')

    opts = get_options()
    tip_len = opts.tip_len

    icol = 3
    for dpath in datpaths:

        model = ofx.Model(dpath)
        model.CalculateStatics()

        all_names = [obj.Name for obj in model.objects]

        stinger_ref = model['b6 stinger_ref']

        radius = float(stinger_ref.tags['radius']) / 1000
        num_section = int(stinger_ref.tags['num_section'])
        st_angles = [-1, -1, -1]
        for isec in range(3):
            st_name = f'b6 stinger_{isec + 1}'
            if st_name in all_names:
                st_angles[isec] = model[st_name].InitialRotation3

        vname = stinger_ref.Connection
        ovessel = model[vname]
        bollard_pull = ovessel.GlobalAppliedForceX[0]

        # Pipe data
        oltype = [obj for obj in model.objects if
                  obj.typeName == 'Line type'][0]  # First linetype in model

        if oltype.Category == 'General':
            # Concrete coated
            cwc_tags = ['cwc thickness (m)', 'cwc density (t/m^3)',
                        'lining thickness (m)', 'lining density (t/m^3)',
                        'snif', 'stress-strain table']
            if not all(k in oltype.tags for k in cwc_tags):
                outstr = 'Pipe linetype is missing one or more tags: '
                outstr += ', '.join([x for x in cwc_tags])
                print(outstr)
                return
            mass = oltype.MassPerUnitLength
            pipe_od = oltype.StressOD
            pipe_id = oltype.StressID
            if pipe_id == ofx.OrcinaDefaultReal():
                pipe_id = oltype.ID
            coating_thk = float(oltype.tags['cwc thickness (m)'])
            coating_dens = float(oltype.tags['cwc density (t/m^3)'])
            lining_thk = float(oltype.tags['lining thickness (m)'])
            lining_dens = float(oltype.tags['lining density (t/m^3)'])
            cwc = 'Yes'
            snif = float(oltype.tags['snif'])
        elif oltype.Category == 'Homogeneous pipe':
            mass = oltype.MassPerUnitLength
            pipe_od = oltype.OD
            pipe_id = oltype.ID
            coating_thk = oltype.CoatingThickness
            coating_dens = oltype.CoatingMaterialDensity
            lining_thk = oltype.LiningThickness
            lining_dens = oltype.LiningMaterialDensity
            cwc = 'No'
            snif = 1.0
        else:
            print('Linetype category must be "General" or "Homogeneous pipe"')
            return

        ltype = LineType(
            oltype.Name, mass, pipe_od, pipe_id, coating_thk, coating_dens,
            lining_thk, lining_dens, cwc, snif)
        wt_in_air, wt_submerged = ltype.weights(0.0)

        roller_names = [x[3:] for x in all_names if x[:5] in ['b6 BR', 'b6 SR']]
        last_roller = model['b6 ' + roller_names[-1]]
        last_roller_arc = float(last_roller.tags['arc']) / 1000

        line = model['Line1']
        top_tension = float(line.StaticResult('Effective tension', ofx.oeEndA))

        # Layback, scope, gain
        firing_line = model['b6 firing_line']
        beadstall = float(firing_line.tags['bead stall'])
        bstall_x = firing_line.StaticResult('X', ofx.oeBuoy(beadstall, 0, 0))
        tdp_x = line.StaticResult('X', ofx.oeTouchdown)
        enda_x = line.StaticResult('X', ofx.oeEndA)
        tdp_arc = line.StaticResult('Arc length', ofx.oeTouchdown)
        layback = float(bstall_x - tdp_x)
        scope_length = float(tdp_arc + (bstall_x - enda_x))
        pipe_gain = scope_length - layback

        # Tip clearance
        clear = 0
        for isup in [1, 2]:
            objx = ofx.oeSupport(isup, 'Line1')
            clear += last_roller.StaticResult(
                'Support contact clearance', objx)
        tip_clearance = float(clear) / 2

        # Static results
        last_section = model[f'b6 stinger_{num_section}']
        vertx = min(last_section.VertexX)
        objx = ofx.oeBuoy(vertx, 0, 0)
        draft = float(last_section.StaticResult('Z', objx)) * -1
        depth = model.environment.WaterDepth
        seabed_clearance = depth - draft

        # PIPE STRESS, STRAIN and CODE CHECKS -----------------------------
        codechecks = model['Code checks']
        oltype.DNVSTF101AlphaPm = 1.0
        youngs_mod = oltype.DNVSTF101E

        line_length = line.CumulativeLength[-1]

        arc_ob = ofx.arSpecifiedArclengths(20.0, last_roller_arc - tip_len)
        arc_st = ofx.arSpecifiedArclengths(last_roller_arc - tip_len,
                                        last_roller_arc + tip_len)
        arc_sb = ofx.arSpecifiedArclengths(last_roller_arc + tip_len,
                                        line_length)

        # Stress and strain
        stress_res = []
        var = 'Max pipelay von Mises strain'
        for arc in [arc_ob, arc_st, arc_sb]:
            if snif > 1.0:
                bend_strain = line.RangeGraph(
                    'Max bending strain', ofx.pnStaticState, None, arc).Mean
                tensile_strain = line.RangeGraph(
                    'Direct tensile strain', ofx.pnStaticState, None, arc).Mean
                worst_zz_strain = bend_strain * snif + tensile_strain
                hoop_stress = line.RangeGraph(
                    'Worst hoop stress', ofx.pnStaticState, None, arc).Mean
                hoop_strain = hoop_stress / youngs_mod
                mod_vm_strain = np.sqrt(worst_zz_strain**2 + hoop_strain**2 - 
                                        worst_zz_strain * hoop_strain)
                static = mod_vm_strain
            else:
                static = line.RangeGraph(var, ofx.pnStaticState, None, arc).Mean
            # stress_res += [static.max()]

        var = 'Max von Mises stress'
        for arc in [arc_ob, arc_st, arc_sb]:
            static = line.RangeGraph(var, ofx.pnStaticState, None, arc).Mean
            stress_res += [static.max()]

        var = 'Max bending strain'
        static = line.RangeGraph(var, ofx.pnStaticState, None, arc_ob).Mean
        stress_res += [static.max()]

        # Code checks
        # Functional / Environmental
        # a -> 1.2 / 0.7
        # b -> 1.1 / 1.3
        f101_res = []
        for gamma_f, gamma_e in [[1.2, 0.7], [1.1, 1.3]]:
            codechecks.DNVSTF101GammaF = gamma_f
            codechecks.DNVSTF101GammaE = gamma_e
            var = 'DNV ST F101 disp. controlled'
            codechecks.DNVSTF101GammaC = 0.8
            static = line.RangeGraph(var, ofx.pnStaticState, None, arc_ob).Mean
            f101_res += [static.max()]
            var = 'DNV ST F101 load controlled'
            for arc, gamma_c in [[arc_ob, 0.8], [arc_st, 1.0], [arc_sb, 1.0]]:
                codechecks.DNVSTF101GammaC = gamma_c
                static = line.RangeGraph(var, ofx.pnStaticState, None, arc).Mean
                f101_res += [static.max()]

        # Rollers
        support_loads = []
        roller_depths = []
        for rname in roller_names:
            oroller = model[f'b6 {rname}']
            nsup = oroller.NumberOfSupports
            react = 0
            for isup in range(nsup):
                objx = ofx.oeSupport(isup + 1)
                react += oroller.StaticResult('Support reaction force', objx)
            support_loads.append(react)
            roller_depths.append(oroller.StaticResult('Z'))

        # Pipe angles
        rnames = ['SR1']
        nroller = len(roller_names)
        for irlr in range(nroller):
            sup_load = support_loads[nroller - 1 - irlr]
            if sup_load > 1.0:
                rnames.append(roller_names[nroller - 1 - irlr])
                break
        pipe_angles = []
        for rname in rnames:
            oroller = model[f'b6 {rname}']
            roller_arc = float(oroller.tags['arc']) / 1000
            pipe_angles.append(line.StaticResult(
                'Declination', ofx.oeArcLength(roller_arc)) - 90.0)

        ws.cell(2, icol).value = dpath.stem
        ws.cell(3, icol).value = ' '.join(vname.split()[1:])
        ws.cell(4, icol).value = model.environment.WaterDepth

        ws.cell(7, icol).value = radius
        ws.cell(8, icol).value = num_section
        irow = 9
        for ang in st_angles:
            if ang > -1:
                ws.cell(irow, icol).value = ang
            irow += 1

        # Pipe data
        istart = 13
        ws.cell(istart + 1, icol).value = ltype.OD * 1000
        ws.cell(istart + 2, icol).value = ltype.WallThickness * 1000
        ws.cell(istart + 3, icol).value = ltype.CWC
        ws.cell(istart + 4, icol).value = ltype.SNIF
        if ltype.CoatingThickness > 0.0:
            ws.cell(istart + 5, icol).value = ltype.CoatingThickness * 1000
            ws.cell(istart + 6, icol).value = ltype.CoatingMaterialDensity * 1000
        if ltype.LiningThickness > 0.0:
            ws.cell(istart + 7, icol).value = ltype.LiningThickness * 1000
            ws.cell(istart + 8, icol).value = ltype.LiningMaterialDensity * 1000
        ws.cell(istart + 9, icol).value = wt_in_air * 1000
        ws.cell(istart + 10, icol).value = wt_submerged * 1000

        # Static results
        istart = 25
        ws.cell(istart + 1, icol).value = bollard_pull
        ws.cell(istart + 2, icol).value = top_tension
        ws.cell(istart + 3, icol).value = layback
        ws.cell(istart + 4, icol).value = scope_length
        ws.cell(istart + 5, icol).value = pipe_gain
        ws.cell(istart + 6, icol).value = tip_clearance
        ws.cell(istart + 7, icol).value = draft
        ws.cell(istart + 8, icol).value = seabed_clearance
        ws.cell(istart + 9, icol).value = pipe_angles[0]
        ws.cell(istart + 10, icol).value = pipe_angles[1]

        # Pipe stress and strain
        for irow in range(7):
            val = stress_res[irow]
            if irow in [3, 4, 5]:
                val /= 1000  # Convert stresses to MPa
            ws.cell(38 + irow, icol).value = val

        # F101 code checks
        for irow in range(8):
            ws.cell(47 + irow, icol).value = f101_res[irow]

        # Support loads
        irow = 57
        for load in support_loads:
            ws.cell(irow, icol).value = load
            irow += 1

        # Roller depths
        irow = 73
        for depth in roller_depths:
            ws.cell(irow, icol).value = depth
            irow += 1

        # Set cell alignment
        for irow in range(1, 100):
            ws.cell(irow, icol).alignment = Alignment(
                horizontal='center', vertical='center')

        # ws.cell(17, icol).number_format = '0.0'
        # ws.cell(18, icol).number_format = '0.0'

        icol += 1

    wb.save(outpath)


def fetch_files():
    """Copy input files into workspace directory:
    *pipe.dat*, *options.xlsx* and *environment.xlsx*.
    """

    for fname in ['pipe.dat', 'options.xlsx', 'environment.xlsx']:
        inpath = files('pypelay') / fname
        print(inpath)
        outpath = PATH / fname
        if outpath.exists():
            inp = input(f'File {fname} already exists. Overwrite (y/n)? ')
            if inp.lower() == 'n':
                return

        shutil.copy(inpath, outpath.resolve())


def set_radius(vessel: Vessel, num_section: int, radius: float,
               water_depth: float, tip_clearance: float,
               outpath: Path) -> None:
    """Create new dat files for specified vessel and stinger configuration.

    Two dat files are created: one with pivoting rollers, one (with *_dyn* suffix)
    with fixed rollers for use in dynamic analysis.

    Pipe segmentation and deadband options are specified in *options.xlsx*.

    Args:
        vessel (Vessel): Vessel object
        num_section (int): Stinger number of sections
        radius (float): Stinger radius in m
        water_depth (float): Water depth in m
        tip_clearance (float): Pipe-roller clearance at stinger tip in m
        outpath (Path): File path of new dat file
    """

    radius *= 1000
    xlpath = files('pypelay') / f'{vessel.name}_configs.xlsx'
    df = pd.read_excel(xlpath)

    df = df[(df['num_section'] == num_section) & 
            (df['prefer'] == 1)]

    df.sort_values(['radius'], inplace=True)
    df.reset_index(drop=True, inplace=True)

    # Find closest radius
    ind = (df['radius'] - radius).abs().argmin()
    config = df.iloc[ind].to_dict()

    model = get_base_case(vessel, config['radius'], num_section)

    stargs = StingerSetupArgs(
        PATH / 'base case.dat', outpath, config, water_depth,
        tip_clearance, delete_dat=False
    )

    res = stinger_setup(stargs)

    (PATH / 'base case.dat').unlink()


def select_radius(vessel: Vessel, num_section: int,
                  water_depth: float, tip_clearance: float,
                  lcc_target: float):
    """Determine optimum stinger radius and create corresponding dat file.

    Two dat files are created: one with pivoting rollers, one (with *_dyn* suffix)
    with fixed rollers for use in dynamic analysis.

    Pipe segmentation and deadband options are specified in *options.xlsx*.

    Args:
        vessel (Vessel): Vessel object
        num_section (int): Stinger number of sections
        water_depth (float): Water depth in m
        tip_clearance (float): Pipe-roller clearance at stinger tip in m
        lcc_target (float): Target F101 LCC utilization in the sag bend
    """

    # STAGE 1: Use simple catenary model to get first estimate of
    # stinger radius
    vmodel = ofx.Model()
    ovessel = vmodel.CreateObject(ofx.ObjectType.Vessel, name='Vessel1')
    vmodel.environment.Depth = water_depth

    # Fetch pipe linetype from pipe dat file
    datpath = PATH / 'pipe.dat'
    pmodel = ofx.Model(datpath)  # Pipe model
    linetypes = [obj.Name for obj in pmodel.objects if 
                    obj.typeName == 'Line type']
    linetype = pmodel[linetypes[0]]
    # Check for stress-strain curve
    if isinstance(linetype.E, str):
        stress_strain = pmodel[linetype.E]
        clone = stress_strain.CreateClone(name=stress_strain.Name,
                                          model=vmodel)
    # Check for non-linear bending stiffness
    if isinstance(linetype.EIx, str):
        bend_stiff = pmodel[linetype.EIx]
        clone = bend_stiff.CreateClone(name=bend_stiff.Name, model=vmodel)
    clone = linetype.CreateClone(name=linetype.Name, model=vmodel)
    line = vmodel.CreateObject(ofx.ObjectType.Line, name='Line1')

    code_checks = vmodel['Code checks']
    code_checks.DNVSTF101GammaF = 1.2
    code_checks.DNVSTF101GammaE = 0.7
    code_checks.DNVSTF101GammaC = 1.0

    line.EndAConnection = 'Vessel1'
    line.EndAX = 0.0
    line.EndAY = 0.0
    line.EndAZ = 0.0
    line.LayAzimuth = 0.0
    line.TargetSegmentLength[0] = 5.0

    # First estimate of top angle used to set line length
    # Should use a formula instead, i.e. top_angle = f(water_depth)
    if water_depth < 100:
        top_angle = math.radians(30.0)
    elif 100 <= water_depth <= 500:
        top_angle = math.radians(50.0)
    else:
        top_angle = math.radians(70.0)

    dh = water_depth
    susp_len, dx = catenary_length(dh, top_angle)
    length_on_seabed = 200
    line.EndBConnection = 'Anchored'
    line.EndBX = -(dx + length_on_seabed)
    line.EndBHeightAboveSeabed = 0.0
    line.Length[0] = susp_len + length_on_seabed

    vmodel.CalculateStatics()

    lcc0 = line.RangeGraph('DNV ST F101 load controlled').Mean.max()

    bp0 = -ovessel.StaticResult('Connections GX force')
    if bp0 < 10:
        bp0 = 10
    bp1 = bp0 * 1.05

    ovessel.IncludedInStatics = '3 DOF'
    ovessel.IncludeAppliedLoads = 'Yes'
    ovessel.NumberOfGlobalAppliedLoads = 1
    ovessel.GlobalAppliedLoadOriginX[0] = 50.0
    ovessel.GlobalAppliedForceX[0] = bp1

    # Iterate bollard pull to get target sag bend UC
    lcc1 = 0.0
    for _ in range(10):
        ovessel.GlobalAppliedForceX[0] = bp1
        vmodel.CalculateStatics()
        lcc1 = line.RangeGraph('DNV ST F101 load controlled').Mean.max()
        if abs(lcc1 - lcc_target) < 0.005:
            break
        bp2 = (lcc_target - lcc0) / (lcc1 - lcc0) * (bp1 - bp0) + bp0
        bp2 = min([float(x) for x in [bp1 * 2, bp2]])
        bp2 = max([float(x) for x in [bp1 * 0.5, bp2]])
        bp0, bp1 = bp1, bp2
        lcc0 = lcc1

    arc_ang = line.RangeGraphXaxis('Declination')
    line_angles = line.RangeGraph('Declination').Mean + 90
    arc_dep = line.RangeGraphXaxis('Z')
    line_depths = line.RangeGraph('Z').Mean
    line_angles = np.interp(arc_dep, arc_ang, line_angles)

    xlpath = files('pypelay') / (f'{vessel.name}_configs.xlsx')
    df = pd.read_excel(xlpath)

    df = df[(df['num_section'] == num_section) & 
            (df['prefer'] == 1)]

    df.sort_values(['radius'], inplace=True)
    df.reset_index(drop=True, inplace=True)

    # Stinger radius is determined from line angle at stinger tip
    # Need to iterate since stinger tip angle varies with tip depth
    radii = df[f'radius'].to_numpy()
    tip_angles = df[f'tip_angle'].to_numpy()
    tip_depths = df[f'tip_depth'].to_numpy()

    top_angle = line_angles[0]
    for _ in range(5):
        rad_mm = float(np.interp(-top_angle, -tip_angles, radii))
        tip_depth = float(np.interp(rad_mm, radii, tip_depths))
        top_angle = float(np.interp(-tip_depth, -line_depths, line_angles))

    ind = (df['radius'] - rad_mm).abs().argmin()

    lc = 1
    toggle = 0
    radii = [0, 0]
    for i in range(5):
        config = df.iloc[ind].to_dict()
        radius = config['radius']

        model = get_base_case(vessel, radius, num_section)

        outpath = PATH / f'tmp{toggle}.dat'
        radii[toggle] = radius
        toggle = 1 - toggle

        stargs = StingerSetupArgs(
            PATH / 'base case.dat', outpath, config, water_depth,
            tip_clearance, delete_dat=False
        )

        res = stinger_setup(stargs)

        (PATH / 'base case.dat').unlink()

        lcc1 = res.uc_sag 
        if lcc1 > lcc_target:
            ind += 1
        else:
            ind += -1

        if i > 0:
            if min(lcc0, lcc1) < lcc_target < max(lcc0, lcc1):
                to_keep = toggle if lcc0 < lcc1 else 1 - toggle
                to_del = 1 - to_keep
                # Delete tmp files not needed
                (PATH / f'tmp{to_del}.dat').unlink()
                (PATH / f'tmp{to_del}_dyn.dat').unlink()

                keep_path = PATH / f'R{radii[to_keep] / 1000:.0f}.dat'
                keep_path_dyn = PATH / f'R{radii[to_keep] / 1000:.0f}_dyn.dat'

                if keep_path.exists():
                    inp = input(f'File {keep_path.name} already exists. Overwrite (y/n)? ')
                    if inp.lower() == 'y':
                        keep_path.unlink()
                        if keep_path_dyn.exists():
                            keep_path_dyn.unlink()
                    else:
                        new_name = input(f'Enter new file name (excluding .dat extension): ')
                        keep_path = PATH / f'{new_name}.dat'
                        keep_path_dyn = PATH / f'{new_name}_dyn.dat'

                (PATH / f'tmp{to_keep}.dat').rename(keep_path)
                (PATH / f'tmp{to_keep}_dyn.dat').rename(keep_path_dyn)
                break

        lcc0 = lcc1
        lc += 1


def stinger_tip_data(model: ofx.Model) -> tuple[float, float, float]:

    stinger_ref = model['b6 stinger_ref']
    num_section = int(stinger_ref.tags['num_section'])

    last_section = model[f'b6 stinger_{num_section}']
    all_names = [obj.Name for obj in model.objects]
    roller_names = [x[3:] for x in all_names if x[:5] in ['b6 BR', 'b6 SR']]
    last_roller = model['b6 ' + roller_names[-1]]

    model.CalculateStatics()

    vertx = min(last_section.VertexX)
    objx = ofx.oeBuoy(vertx, 0, 0)
    draft = float(last_section.StaticResult('Z', objx))
    # tip angle is equal to stinger section angle plus roller r3
    conn = last_section.Connection
    last_section.Connection = 'Fixed'
    section_angle = float(last_section.InitialRotation3 + 180)
    last_section.Connection = conn

    conn = last_roller.Connection
    last_roller.Connection = 'Fixed'
    tip_depth = float(last_roller.InitialZ)
    last_roller.Connection = conn

    roller_r3 = float(last_roller.tags['r3'])
    tip_angle = section_angle + roller_r3

    model.CalculateStatics()

    return tip_depth, tip_angle, draft


def get_options() -> StingerSetupOptions:

    # Ignore Data validation warning
    with warnings.catch_warnings(action='ignore', category=UserWarning):
        df = pd.read_excel(PATH / 'options.xlsx', sheet_name='Sheet1')

    df.set_index('ID', inplace=True)
    opts = StingerSetupOptions(
        tip_def = df.loc[1, 'Value'],
        tip_len = df.loc[2, 'Value'],
        seglen = df.loc[3, 'Value'],
        ob_seglen = df.loc[4, 'Value'],
        tdp_seglen = df.loc[5, 'Value'],
        tdp_len = df.loc[6, 'Value'],
        tensioner_mode = df.loc[7, 'Value'],
        deadband = df.loc[8, 'Value'],
    )

    return opts


def stinger_setup(sim: StingerSetupArgs) -> StingerSetupResults:
    '''
     - Calculate pipe path and roller heights
     - Set roller heights
     - Create pipe
     - Adjust bollard pull to get target tip clearance
     - Update segmentation around TDP
     - Add deadband winch if needed
    '''
    model = ofx.Model(sim.inpath)
    # print(sim.outpath)

    # Read options spreadsheet
    opts = get_options()

    model.environment.WaterDepth = sim.water_depth

    ang1 = sim.config['ang1']
    ang2 = sim.config['ang2']
    straight = sim.config['straight']
    transition = sim.config['transition']

    # Set stinger section angles
    stinger_ref = model['b6 stinger_ref']
    num_section = int(stinger_ref.tags['num_section'])
    radius = int(stinger_ref.tags['radius'])
    model['b6 stinger_1'].InitialRotation3 = ang1
    if num_section > 1:
        model['b6 stinger_2'].InitialRotation3 = ang2

    # Set roller heights and angles
    path_coords = calc_path_coords(model, straight, transition)
    rollers = get_roller_heights(model, path_coords, ang1, ang2)

    # np.savetxt(PATH / 'path_coords.txt', path_coords, delimiter=',') xxx

    if not rollers:
        print('Stinger config not valid, roller heights outside limits')
        return StingerSetupResults(sim.outpath)
    
    for roller in rollers:
        oroller = model[f'cn {roller.name}']
        oroller.InitialY = roller.y / 1000
        oroller.DOFInitialValue[5] = roller.r3
        # Manual roller height adjustments
        if roller.name in sim.config:
            manual_roller_y = sim.config[roller.name]
            if np.isnan(manual_roller_y):
                continue
            oroller.InitialY = manual_roller_y / 1000

    # Create line
    roller = rollers[-1]
    # Connect line to roller at y_offset then change connection to stinger_ref
    line = model.CreateObject(ofx.ObjectType.Line, 'Line1')
    line.LayAzimuth = 0.0
    line.NumberOfSections = 3
    # Line overbend (first section) extends 10m past last roller
    ob_length = roller.arc / 1000 + 10.0
    line.Length[0] = 20.0  # First section is for tensioner
    line.TargetSegmentLength[0] = 20.0
    line.Length[1] = ob_length - 20.0
    line.TargetSegmentLength[1] = opts.ob_seglen
    line.TargetSegmentLength[2] = opts.seglen
    line.FullStaticsMinDamping = 5.0
    line.FullStaticsMaxDamping = 20.0

    buoy_name = f'b6 {roller.name}'
    line.EndAConnection = buoy_name
    line.EndAX = 0.0
    line.EndAY = roller.y_offset / 1000
    line.EndAZ = 0.0
    line.EndAConnection = 'Fixed'
    end_a = np.array([line.EndAX, line.EndAZ])
    line.EndAConnection = 'b6 stinger_ref'

    top_angle = math.radians(roller.post_angle + roller.r3 - 90)
    dh = end_a[1] + sim.water_depth
    susp_len, dx = catenary_length(dh, top_angle)
    length_on_seabed = 200
    line.EndBConnection = 'Anchored'
    line.EndBX = end_a[0] - dx - length_on_seabed
    line.EndBHeightAboveSeabed = 0.0

    line.Length[2] = susp_len + length_on_seabed
    line.EndAConnection = 'b6 stinger_ref'
    line.EndAX = path_coords[-1, 0] / 1000
    line.EndAY = path_coords[-1, 1] / 1000
    # Set top end orientation and stiffness
    line.IncludeTorsion = 'Yes'
    line.EndAxBendingStiffness = ofx.OrcinaInfinity()
    line.EndATwistingStiffness = ofx.OrcinaInfinity()
    line.EndAAzimuth = 180
    line.EndADeclination = 90
    line.EndAGamma = 180

    # Set rollers supported line
    model['b6 firing_line'].NumberOfSupportedLines = 1
    model['b6 firing_line'].SupportedLine[0] = 'Line1'
    for roller in rollers:
        buoy = model[f'b6 {roller.name}']
        buoy.NumberOfSupportedLines = 1
        buoy.SupportedLine[0] = 'Line1'
        buoy.tags['post_angle'] = f'{roller.post_angle:.1f}'
        buoy.tags['r3'] = f'{roller.r3:.3f}'
        buoy.tags['arc'] = f'{roller.arc:.1f}'

    # For first solve set rollers to soft
    # set_roller_stiffness(model, 3000)

    # Move anchor for large stinger radii
    # set_roller_pivots(model, option='fixed')
    if radius > 240000:
        line.EndBX += -12

    # Solve statics, use calculated line shapes
    solved = False
    for _ in range(5):
        try:
            model.CalculateStatics()
            model.UseCalculatedPositions(SetLinesToUserSpecifiedStartingShape=True)
            solved = True
            break
        except:
            newdamp = line.FullStaticsMinDamping + 2.0
            line.FullStaticsMinDamping = newdamp
            model.SaveData(sim.outpath)
            print(f'Loadcase {sim.outpath.name} increasing damping to {newdamp}')

    if not solved:
        print(f'Loadcase {sim.outpath.name} failed to solve')
        return StingerSetupResults(sim.outpath)

    # Write stinger config details to b6 stinger_ref tags,
    # add comment to General->Comments
    buoy = model['b6 stinger_ref']
    buoy.tags['date'] = datetime.datetime.now().strftime('%Y-%m-%d')
    # buoy.tags['pipe_od'] = f'{pipe_od:.1f}'
    buoy.tags['straight'] = f'{straight:.1f}'
    buoy.tags['transition'] = f'{transition:.1f}'

    outstr = f'Stinger configuration generated with Python pipelay module.\n'
    outstr += 'For details refer to "b6 stinger_ref" tags.'
    model.general.Comments = outstr

    # Adjust bollard pull to get target tip clearance
    # (updates model, unsolved state)
    solved = set_bollard_pull(model, 'tip_clearance', sim.tip_clearance, tol=0.01)
    if not solved:
        print('Failed to converge (tip clearance)')
        return StingerSetupResults(sim.outpath)

    update_segmentation(model, opts)

    if opts.tensioner_mode != 'Brake':
        add_deadband_winch(model, opts)

    # Increase roller stiffness
    set_roller_stiffness(model, 70e3)

    model.SaveData(sim.outpath)

    # Create dat file suitable for running dynamics (vessel)
    prep_for_dyn(sim.outpath)

    res = get_setup_results(model, sim.outpath)

    # If running all 18000 configs need to delete dat files
    if sim.delete_dat:
        sim.outpath.unlink()

    print(f'Radius {radius / 1000:.0f}m : {res}')

    return res


def prep_for_dyn(datpath: Path) -> None:
    '''
    Prepare model for dynamic analysis:
     - Set roller orientations, delete constraints
     - Switch off vessel 3 DOF (so that current doesn't effect position)
    '''
    model = ofx.Model(datpath)

    set_roller_pivots(model, option='dyn')

    ovessel = [obj for obj in model.objects if obj.typeName == 'Vessel'][0]
    ovessel.IncludedInStatics = 'None'

    outpath = datpath.parent / f'{datpath.stem}_dyn.dat'
    model.SaveData(outpath)


def set_roller_stiffness(model: ofx.Model, stiffness: float) -> None:
    '''Set stiffness of all support types'''
    suptypes = [obj for obj in model.objects if 
                obj.type == ofx.ObjectType.SupportType]
    for stype in suptypes:
        stype.NormalStiffness = stiffness


def set_roller_pivots(model: ofx.Model, option: str = 'dyn') -> None:
    '''
    Options:
    1. fixed, release: Fix all roller pivots to theoretical angles
    2. released: Release all rollers, set initial angles to theoretical
    2. dyn: Fix roller positions for dynamic analysis
    '''
    all_names = [obj.Name for obj in model.objects]

    roller_names = [x[3:] for x in all_names if x[:5] in ['b6 BR', 'b6 SR']]

    match option:
        case 'fixed':
            for rname in roller_names:
                oroller = model[f'b6 {rname}']
                ocn = model[f'cn {rname}']
                ocn.DOFFree[5] = 'No'
                r3 = float(oroller.tags['r3'])
                oroller.InitialRotation3 = r3

        case 'released':
            for rname in roller_names:
                oroller = model[f'b6 {rname}']
                ocn = model[f'cn {rname}']
                ocn.DOFFree[5] = 'Yes'
                r3 = float(oroller.tags['r3'])
                ocn.DOFInitialValue[5] = r3
                oroller.InitialRotation3 = 0.0

        case 'dyn':
            # Determine if roller angle needs to be set to r3 based on
            # roller load
            model.CalculateStatics()
            adjust_angle = {}
            for rname in roller_names:
                adjust_angle[rname] = False
                if option == 'dyn':
                    oroller = model[f'b6 {rname}']
                    nsup = oroller.NumberOfSupports
                    for isup in range(nsup):
                        objx = ofx.oeSupport(isup + 1)
                        sup_load = oroller.StaticResult('Support reaction force', objx)
                        if sup_load < 0.1:
                            adjust_angle[rname] = True
                            break

            for rname in roller_names:
                oroller = model[f'b6 {rname}']
                ocn = model[f'cn {rname}']
                oroller.Connection = ocn.Connection
                if adjust_angle[rname]:
                    r3 = float(oroller.tags['r3'])
                    oroller.InitialRotation3 = r3
                model.DestroyObject(ocn)

 
def add_deadband_winch(model: ofx.Model, opts: StingerSetupOptions) -> None:

    if model.state == 0:
        # Reset state = 0, InStaticState = 2
        model.CalculateStatics()

    line = model['Line1']
    top_tension = line.StaticResult('Effective tension', ofx.oeEndA)
    match opts.tensioner_mode:
        case 'Deadband: Default':
            if top_tension <= 100 * 9.81:
                deadband = 10 * 9.81
            else:
                deadband = top_tension * 0.1
        case 'Deadband: % top tension':
            deadband = top_tension * opts.deadband
        case 'Deadband: tension in t':
            deadband = opts.deadband * 9.81

    # Create tensioner linetype
    lt_tens = model.CreateObject(ofx.ObjectType.LineType, name='lt tensioner')
    lt_tens.Category = 'General'
    lt_pipe = model[line.LineType[0]]
    lt_tens.OD = lt_pipe.OuterContactDiameter
    lt_tens.ID = lt_pipe.ID
    lt_tens.MassPerUnitLength = lt_pipe.MassPerUnitLength
    lt_tens.EIx = 200E3
    lt_tens.GJ = 140E3
    lt_tens.EA = 1.0
    line.Linetype[0] = 'lt tensioner'

    # Create tensioner winch, both ends connected to pipe
    winch = model.CreateObject(ofx.ObjectType.Winch, name='w tensioner')
    winch.WinchType = 'Detailed'
    for iconn in [0, 1]:
        winch.Connection[iconn] = 'Line1'
        for dof in ['X', 'Y', 'Z']:
            winch.SetData(f'Connection{dof}', iconn, 0.0)
    #     winch.ConnectionZ[0] = 0.0
    # winch.Connection[1] = 'Line1'
    winch.ConnectionZ[1] = 20.0

    winch.WinchControlType = 'By stage'
    winch.StageMode[0] = 'Specified length'
    winch.StageValue[0] = 20.0
    for istg in [1, 2]:
        winch.StageMode[istg] = 'Specified tension change'
        winch.StageValue[istg] = 0.0

    winch.Stiffness = 100E3
    winch.Damping = 0.4
    winch.DriveDeadband = deadband
    winch.DriveDampingHaulIn = 0.1
    winch.DriveDampingPayOut = 0.1


def update_segmentation(model: ofx.Model, opts: StingerSetupOptions) -> None:

    if model.state == 0:
        # Reset state = 0, InStaticState = 2
        model.CalculateStatics()

    line = model['Line1']

    tdp_arc = line.StaticResult('Arc length', ofx.oeTouchdown)

    tdp_len = opts.tdp_len

    # Section 1 is 20m single segment for tensioner winch
    # Section 2 is overbend, from start of line to 10m past last roller
    # (does not change)
    # Section 3 is up to TDP section
    # Section 4 is TDP region
    # Section 5 is from TDP region to end of line

    len0 = line.CumulativeLength[1]
    line_length = line.CumulativeLength[-1]

    extends_to_tip = False
    if (tdp_arc - tdp_len / 2) < (len0 + 2):
        extends_to_tip = True

    extends_to_end = False
    if tdp_arc + tdp_len / 2 > line_length - 5:
        extends_to_end = True

    if all([extends_to_tip, extends_to_end]):
        line.NumberOfSections = 3
        line.Length[2] = line_length - len0
        line.TargetSegmentLength[2] = opts.tdp_seglen

    if all([extends_to_tip, not extends_to_end]):
        line.NumberOfSections = 4
        line.Length[2] = tdp_arc + tdp_len / 2 - len0
        line.TargetSegmentLength[2] = opts.tdp_seglen
        line.Length[3] = line_length - (tdp_arc + tdp_len / 2)
        line.TargetSegmentLength[3] = opts.seglen

    if all([not extends_to_tip, extends_to_end]):
        line.NumberOfSections = 4
        line.Length[2] = tdp_arc - tdp_len / 2 - len0
        line.TargetSegmentLength[2] = opts.seglen
        line.Length[3] = line_length - (tdp_arc - tdp_len / 2)
        line.TargetSegmentLength[3] = opts.tdp_seglen

    if all([not extends_to_tip, not extends_to_end]):
        line.NumberOfSections = 5
        line.Length[2] = tdp_arc - tdp_len / 2 - len0
        line.TargetSegmentLength[2] = opts.seglen
        line.Length[3] = tdp_len
        line.TargetSegmentLength[3] = opts.tdp_seglen
        line.Length[4] = line_length - (tdp_arc + tdp_len / 2)
        line.TargetSegmentLength[4] = opts.seglen

    line.StaticsStep1 = 'Catenary'

    solved = False
    for _ in range(5):
        try:
            model.CalculateStatics()
            model.UseCalculatedPositions(SetLinesToUserSpecifiedStartingShape=True)
            solved = True
            break
        except:
            newdamp = line.FullStaticsMinDamping + 2.0
            line.FullStaticsMinDamping = newdamp
            print(f'Increasing line damping to {newdamp}')

    if not solved:
        print('Couldnt solve after updating segmentation')


def adjust_top_tension(inpath: Path, outpath: Path, tension: float) -> None:
    """Adjust top tension for an existing stinger configuration.

    Args:
        inpath (Path): File path of existing dat file
        outpath (Path): File path of new dat file
        tension (float): Target top tension in kN
    """

    model = ofx.Model(inpath)

    model.SaveData(outpath)

    solved = set_bollard_pull(model, 'top_tension', tension, tol=1.0)

    if not solved:
        print('Failed to converge (top tension)')
        return

    opts = get_options()
    update_segmentation(model, opts)

    model.SaveData(outpath)

    prep_for_dyn(outpath)

    res = get_setup_results(model, outpath)

    print(res)


def set_bollard_pull(model: ofx.Model, 
                     target_var: str, target_val: float,
                     tol: float) -> bool:
    # Iterates vessel bollard pull to achieve target stinger tip clearance
    # model = ofx.Model(datpath)

    all_names = [obj.Name for obj in model.objects]
    roller_names = [x[3:] for x in all_names if x[:5] in ['b6 BR', 'b6 SR']]
    oroller = model['b6 ' + roller_names[-1]]

    line = model['Line1']
    vessel = [obj for obj in model.objects if obj.typeName == 'Vessel'][0]

    model.CalculateStatics()

    if target_var == 'tip_clearance':
        clr_sup1 = oroller.StaticResult('Support contact clearance', ofx.oeSupport(1))
        clr_sup2 = oroller.StaticResult('Support contact clearance', ofx.oeSupport(2))
        res0 = (clr_sup1 + clr_sup2) / 2
    else:
        res0 = line.StaticResult('Effective tension', ofx.oeEndA)

    bp0 = -vessel.StaticResult('Connections GX force')
    if bp0 < 10:
        bp0 = 10
    bp1 = bp0 * 1.05
    vessel.IncludedInStatics = '3 DOF'

    for _ in range(10):
        vessel.GlobalAppliedForceX[0] = bp1
        # Solve statics, use calculated line shapes
        solved = False
        for _ in range(5):
            try:
                model.CalculateStatics()
                model.UseCalculatedPositions(SetLinesToUserSpecifiedStartingShape=True)
                solved = True
                break
            except:
                newdamp = line.FullStaticsMinDamping + 2.0
                line.FullStaticsMinDamping = newdamp
                print(f'Increasing line damping to {newdamp}')

        if not solved:
            return False

        model.CalculateStatics()

        if target_var == 'tip_clearance':
            clr_sup1 = oroller.StaticResult('Support contact clearance', ofx.oeSupport(1))
            clr_sup2 = oroller.StaticResult('Support contact clearance', ofx.oeSupport(2))
            res1 = (clr_sup1 + clr_sup2) / 2
        else:
            res1 = line.StaticResult('Effective tension', ofx.oeEndA)

        if abs(res1 - target_val) < tol:
            break

        bp2 = (target_val - res0) / (res1 - res0) * (bp1 - bp0) + bp0
        # Bound the new BP to avoid too large jump
        bp2 = max(bp1 / 2, min(bp1 * 2, bp2))
        # print(bp0, bp1, bp2)
        # bp2 = min([float(x) for x in [bp1 * 2, bp2]])
        bp0, bp1 = bp1, bp2
        res0 = res1

    model.UseCalculatedPositions(SetLinesToUserSpecifiedStartingShape=True)

    return True


def get_setup_results(model: ofx.Model, datpath: Path) -> StingerSetupResults:

    if model.state == 0:
        # Reset state = 0, InStaticState = 2
        model.CalculateStatics()

    all_names = [obj.Name for obj in model.objects]
    roller_names = [x[3:] for x in all_names if x[:5] in ['b6 BR', 'b6 SR']]
    oroller = model['b6 ' + roller_names[-1]]

    line = model['Line1']
    vessel = [obj for obj in model.objects if obj.typeName == 'Vessel'][0]

    tip_depth, tip_angle, draft = stinger_tip_data(model)

    code_checks = model['Code checks']
    code_checks.DNVSTF101GammaF = 1.2
    code_checks.DNVSTF101GammaE = 0.7

    top_tension = float(line.StaticResult('Effective tension', ofx.oeEndA))

    last_roller_arc = float(oroller.tags['arc']) / 1000
    line_length = line.CumulativeLength[-1]
    arc_ob = ofx.arSpecifiedArclengths(20.0, last_roller_arc)
    arc_sag = ofx.arSpecifiedArclengths(last_roller_arc, line_length)

    var = 'DNV ST F101 load controlled'
    code_checks.DNVSTF101GammaC = 0.8
    uc_ob = line.RangeGraph(var, None, None, arc_ob).Mean.max()
    code_checks.DNVSTF101GammaC = 1.0
    uc_sag = line.RangeGraph(var, None, None, arc_sag).Mean.max()

    return StingerSetupResults(datpath, top_tension, uc_ob, uc_sag,
                               tip_depth, tip_angle, draft)


def remove_stinger_section(model: ofx.Model, isect: int):
    ''' Removes stinger section and associated constraints and rollers'''
    all_names = [obj.Name for obj in model.objects]

    constraints = [obj for obj in model.objects if obj.typeName == 'Constraint']
    for cn in constraints:
        if cn.Name[:5] == 'cn SR':
            if cn.InFrameConnection == f'b6 stinger_{isect}':
                buoy = cn.Name.replace('cn', 'b6')
                model.DestroyObject(cn)
                model.DestroyObject(buoy)

    model.DestroyObject(model[f'b6 stinger_{isect}'])

    return model


def get_base_case(vessel: Vessel, stinger_radius: float,
                  num_section: int = 3) -> ofx.Model:
    # - Add pipe linetype to vessel model,
    # - Delete stinger sections,
    # - update stinger_ref tags (radius, num_section, path length)

    datpath = Path(str(files('pypelay') / f'{vessel.name}.dat'))
    model = ofx.Model(datpath)

    # Set vessel type and draft
    ovessel = model[f'v {vessel.name}']
    ovessel.VesselType = vessel.vessel_type
    ovessel.Draught = vessel.draft_name
    ovessel.InitialZ = -vessel.draft

    # Fetch pipe linetype from pipe dat file
    datpath = PATH / 'pipe.dat'
    pmodel = ofx.Model(datpath)
    linetypes = [obj.Name for obj in pmodel.objects if 
                 obj.typeName == 'Line type']
    linetype = pmodel[linetypes[0]]
    # Category Homogeneous pipe: Check for stress-strain curve
    if isinstance(linetype.E, str):
        stress_strain = pmodel[linetype.E]
        clone = stress_strain.CreateClone(
            name=stress_strain.Name, model=model)
    # Category General: Check for non-linear bending stiffness
    if isinstance(linetype.EIx, str):
        bend_stiff = pmodel[linetype.EIx]
        clone = bend_stiff.CreateClone(name=bend_stiff.Name, model=model)
    # Category General: Stress-strain table is specified in linetype tags
    if 'stress-strain table' in linetype.tags:
        try:
            stress_strain = pmodel[linetype.tags['stress-strain table']]
            clone = stress_strain.CreateClone(name=stress_strain.Name,
                                              model=model)
        except KeyError:
            print(('Stress-strain table specified in linetype tags '
                   'not found in model'))
    else:
        print('Linetype must contain tag "stress-strain table" ')

    clone = linetype.CreateClone(name=linetype.Name, model=model)

    # Remove stinger sections
    if num_section < 3:
        model = remove_stinger_section(model, 3)
    if num_section < 2:
        model = remove_stinger_section(model, 2)

    # Update stinger_ref tags
    stinger_ref = model['b6 stinger_ref']
    stinger_ref.tags['radius'] = f'{stinger_radius:.0f}'
    stinger_ref.tags['num_section'] = f'{num_section}'
    pthlen = stinger_ref.tags['path_length'].split('\n')
    pthlen = pthlen[num_section - 1].split(':')[1].strip()
    stinger_ref.tags['path_length'] = pthlen

    datpath = PATH / 'base case.dat'
    model.SaveData(datpath)

    return model


def get_roller_heights(
    model: ofx.Model, path_coords: np.ndarray,
    ang1: float, ang2: float) -> list[Roller]:
    ''' Calculates roller heights and angles for input stinger angles and pipe path.
        If pipe path is outside roller range (ymin, ymax) then returns None.'''

    all_names = [obj.Name for obj in model.objects]

    stinger_ref = model['b6 stinger_ref']
    radius = int(stinger_ref.tags['radius'])

    # Set stinger angles
    model['b6 stinger_1'].InitialRotation3 = ang1
    if 'b6 stinger_2' in all_names:
        model['b6 stinger_2'].InitialRotation3 = ang2

    ltypes = [obj for obj in model.objects if obj.typeName == 'Line type']
    pipe_od = ltypes[0].OD * 1000

    roller_names = [x[3:] for x in all_names if x[:5] in ['b6 BR', 'b6 SR']]
    # roller_names = [x[3:] for x in all_names if x[:5] in ['b6 SR']] xxx
    nroller = len(roller_names)

    rollers = []
    for iroller, rname in enumerate(roller_names):
        # Get roller data from Orcaflex dat file
        buoy = model[f'b6 {rname}']
        cn = model[f'cn {rname}']
        # roller_conn = cn.InFrameConnection
        cn.InitialY = 0.0
        cn.DOFInitialValue[5] = 0.0
        buoy.Connection = 'b6 stinger_ref'

        roller_data = Roller(rname)
        origin_x = buoy.InitialX * 1000
        origin_y = buoy.InitialY * 1000
        post_angle = buoy.InitialRotation3 + 90
        ymin = float(buoy.tags['ymin']) * 1000
        ymax = float(buoy.tags['ymax']) * 1000
        yincr = float(buoy.tags['yincr']) * 1000
        length = float(buoy.tags['length']) * 1000
        v_angle = float(buoy.tags['v_angle'])
        # Very important: reset Orcaflex model roller
        buoy.Connection = f'cn {rname}'

        # Reduce max height for last roller to allow space for vertical rollers
        tip_capture = 2100
        if iroller == nroller - 1:
            ymax += -tip_capture

        roller_coords = np.array([origin_x, origin_y], dtype=float)
        # Get pipe path coords relative to roller_coords
        pipe_path2 = path_coords - roller_coords
        # Rotate pipe path so that support post is Y-axis
        c = math.cos(math.radians(post_angle - 90))
        s = math.sin(math.radians(post_angle - 90))
        xdash = pipe_path2[:, 0] * c + pipe_path2[:, 1] * s
        ydash = -pipe_path2[:, 0] * s + pipe_path2[:, 1] * c

        # Get pipe slope and curvature
        ind = np.argmin(np.abs(xdash))
        if xdash[ind] > 0:
            dx, dy = path_coords[ind + 1, :] - path_coords[ind, :]
        else:
            dx, dy = path_coords[ind, :] - path_coords[ind - 1, :]
        slope = math.degrees(math.atan2(dy, dx))
        roller_data.post_angle = post_angle
        roller_data.r3 = slope - (post_angle - 90)

        y_offset = (pipe_od/2) / math.cos(math.radians(v_angle))
        # Offset due to pipe curvature and distance between rollers
        curv = menger_curv(path_coords[ind - 1:ind + 2, :])
        if curv > 0:
            R = 1 / curv  # Pipe bend radius at roller
            y_offset += R * (1 - math.cos(length / (2 * R)))

        roller_data.y_offset = y_offset
        y_intersect = np.interp(0, xdash, ydash)
        y_roller = y_intersect - y_offset

        # Calculate pipe_path arc length at roller
        ind = xdash > 0
        xdash = np.insert(xdash[ind], 0, 0.0)
        ydash = np.insert(ydash[ind], 0, y_intersect)

        segdists = np.hypot(np.diff(xdash), np.diff(ydash))
        roller_data.arc = segdists.sum()

        if (ymin - yincr/2) <= y_roller <= (ymax + yincr/2):
            # Adjust roller height to nearest pin hole
            nhole = round((ymax - ymin) / yincr + 1)
            y_hole = np.linspace(ymin, ymax, nhole)
            ind = np.argmin(np.abs(y_hole - y_roller))
            y_roller = y_hole[ind]
            roller_data.y = y_roller
        else:
            return []

        # Corrections for large stinger radii (need to fix this, see TODO)
        # BR5 ymax in dat file is set to 1.25, but should be 1.0
        vname = stinger_ref.Connection
        if vname in ['v S1200', 'v S3500']:
            r1, r2 = (5, 6) if vname == 'v S1200' else (6, 7)
            if radius >= 370000:
                if rname == f'BR{r1}':
                    if roller_data.y > 1425:
                        roller_data.y = 1425
            if rname == f'BR{r2}':
                if roller_data.y > 1000:
                    roller_data.y = 1000

        rollers.append(roller_data)

    return rollers


def main():
    pass


if __name__ == "__main__":
    main()
